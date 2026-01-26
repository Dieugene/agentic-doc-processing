"""
Converter module - document format conversion to PDF.

Supports conversion from DOCX, Excel, TXT to PDF format for subsequent
PNG rendering and VLM-OCR processing.
"""
from __future__ import annotations

import logging
import os
import tempfile
from enum import Enum
from pathlib import Path
from typing import Optional

try:
    from chardet import detect
except ImportError:
    detect = None

try:
    from fpdf import FPDF
except ImportError:
    FPDF = None

try:
    from docx import Document as DocxDocument
except ImportError:
    DocxDocument = None

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None


logger = logging.getLogger(__name__)


class FileType(str, Enum):
    """Supported file types for conversion."""

    PDF = "pdf"
    DOCX = "docx"
    XLSX = "xlsx"
    TXT = "txt"
    UNKNOWN = "unknown"


class ConversionError(Exception):
    """Raised when file conversion fails."""

    def __init__(self, file_type: str, details: str):
        self.file_type = file_type
        self.details = details
        super().__init__(f"Failed to convert {file_type} file: {details}")


class Converter:
    """Converter for various document formats to PDF."""

    # Page dimensions for A4
    A4_WIDTH = 210  # mm
    A4_HEIGHT = 297  # mm
    MARGIN = 15  # mm

    def __init__(self, log_dir: Optional[str] = None):
        """Initialize the Converter.

        Args:
            log_dir: Optional directory for conversion logs.
        """
        self._setup_logging(log_dir)
        self._check_dependencies()

    def _setup_logging(self, log_dir: Optional[str]) -> None:
        """Setup logging for conversion operations."""
        if log_dir:
            log_path = Path(log_dir) / "conversion.log"
            log_path.parent.mkdir(parents=True, exist_ok=True)

            handler = logging.FileHandler(log_path)
            handler.setFormatter(
                logging.Formatter(
                    "%(asctime)s - %(name)s - %(levelname)s - %(message)s"
                )
            )
            logger.addHandler(handler)

    def _check_dependencies(self) -> None:
        """Check if required dependencies are installed."""
        if FPDF is None:
            raise ImportError("fpdf2 is required. Install with: pip install fpdf2")
        if DocxDocument is None:
            raise ImportError("python-docx is required. Install with: pip install python-docx")
        if load_workbook is None:
            raise ImportError("openpyxl is required. Install with: pip install openpyxl")
        if detect is None:
            logger.warning("chardet not installed. TXT encoding detection may fail.")

    async def detect_file_type(self, file_path: str) -> FileType:
        """Detect file type by extension.

        Args:
            file_path: Path to the file

        Returns:
            FileType enum value
        """
        ext = Path(file_path).suffix.lower().lstrip(".")
        try:
            return FileType(ext)
        except ValueError:
            return FileType.UNKNOWN

    async def convert_to_pdf(
        self,
        file_path: str,
        file_type: Optional[FileType] = None,
    ) -> str:
        """Convert a file to PDF format.

        Args:
            file_path: Path to the source file
            file_type: File type (if None, auto-detect)

        Returns:
            Path to the temporary PDF file

        Raises:
            ValueError: If format is not supported
            ConversionError: If conversion fails
            FileNotFoundError: If source file doesn't exist
        """
        # Check if file exists
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File not found: {file_path}")

        # Detect file type if not provided
        if file_type is None:
            file_type = await self.detect_file_type(file_path)

        # Validate file type
        if file_type == FileType.UNKNOWN:
            raise ValueError(
                f"Unsupported file format: {Path(file_path).suffix}. "
                f"Supported: {', '.join([ft.value for ft in FileType if ft != FileType.UNKNOWN])}"
            )

        # Log conversion start
        logger.info(f"Starting conversion: {file_path} ({file_type.value})")

        try:
            # Dispatch to appropriate converter
            if file_type == FileType.PDF:
                # PDF files don't need conversion
                logger.info(f"File is already PDF, returning as-is: {file_path}")
                return file_path

            temp_pdf_path = tempfile.NamedTemporaryFile(
                suffix=".pdf", delete=False
            ).name

            if file_type == FileType.DOCX:
                result_path = await self._convert_docx_to_pdf(file_path, temp_pdf_path)
            elif file_type == FileType.XLSX:
                result_path = await self._convert_xlsx_to_pdf(file_path, temp_pdf_path)
            elif file_type == FileType.TXT:
                result_path = await self._convert_txt_to_pdf(file_path, temp_pdf_path)
            else:
                # This shouldn't happen due to UNKNOWN check above
                raise ValueError(f"Unsupported file type: {file_type}")

            logger.info(f"Conversion successful: {result_path}")
            return result_path

        except Exception as e:
            error_msg = str(e)
            logger.error(f"Conversion failed: {error_msg}")
            raise ConversionError(file_type.value, error_msg) from e

    async def _convert_txt_to_pdf(self, txt_path: str, output_path: str) -> str:
        """Convert plain text file to PDF.

        Args:
            txt_path: Path to source TXT file
            output_path: Path for output PDF file

        Returns:
            Path to the generated PDF file
        """
        # Read file with encoding detection
        text_content = self._read_text_file(txt_path)

        # Create PDF
        pdf = FPDF()
        pdf.set_margin(self.MARGIN)
        pdf.add_page()

        # Use monospace font for text files
        try:
            pdf.set_font("Courier", size=10)
        except Exception:
            # Fallback to default font if Courier not available
            pdf.set_font("Arial", size=10)

        # Add text with line breaks
        line_height = 5  # mm
        page_width = self.A4_WIDTH - 2 * self.MARGIN

        for line in text_content.split("\n"):
            # Handle long lines with wrapping
            if pdf.get_string_width(line) > page_width:
                # Simple word wrapping
                words = line.split(" ")
                current_line = ""
                for word in words:
                    test_line = current_line + " " + word if current_line else word
                    if pdf.get_string_width(test_line) > page_width:
                        if current_line:
                            pdf.cell(0, line_height, txt=current_line, ln=True)
                        current_line = word
                    else:
                        current_line = test_line
                if current_line:
                    pdf.cell(0, line_height, txt=current_line, ln=True)
            else:
                pdf.cell(0, line_height, txt=line, ln=True)

        pdf.output(output_path)
        return output_path

    async def _convert_docx_to_pdf(self, docx_path: str, output_path: str) -> str:
        """Convert DOCX file to PDF.

        Args:
            docx_path: Path to source DOCX file
            output_path: Path for output PDF file

        Returns:
            Path to the generated PDF file
        """
        doc = DocxDocument(docx_path)

        pdf = FPDF()
        pdf.set_margin(self.MARGIN)
        pdf.add_page()

        # Try to use a Unicode-compatible font
        # For v1.0, we use standard fonts. Cyrillic support may be limited.
        try:
            pdf.set_font("Arial", size=12)
        except Exception:
            pdf.set_font("Times", size=12)

        line_height = 7  # mm
        page_width = self.A4_WIDTH - 2 * self.MARGIN

        for para in doc.paragraphs:
            # Extract style information
            bold = False
            italic = False
            font_size = 12

            if para.runs:
                # Use style from first run
                first_run = para.runs[0]
                if first_run.bold:
                    bold = True
                if first_run.italic:
                    italic = True
                if first_run.font.size:
                    font_size = int(first_run.font.size / 2)  # Convert to points (approx)

            # Apply formatting
            pdf.set_font_size(font_size)
            try:
                pdf.set_font("Arial", style="BI" if bold and italic else "B" if bold else "I" if italic else "")
            except Exception:
                pass

            # Get paragraph text
            text = para.text.strip()
            if not text:
                pdf.ln(line_height / 2)
                continue

            # Check if we need a new page
            if pdf.y > self.A4_HEIGHT - 30:
                pdf.add_page()

            # Handle long paragraphs with wrapping
            if pdf.get_string_width(text) > page_width:
                # Simple character wrapping for long lines
                chars_per_line = int(page_width / pdf.get_string_width("x"))
                for i in range(0, len(text), chars_per_line):
                    chunk = text[i:i + chars_per_line]
                    pdf.cell(0, line_height, txt=chunk, ln=True)
            else:
                pdf.cell(0, line_height, txt=text, ln=True)

            # Add spacing after paragraph
            pdf.ln(2)

        pdf.output(output_path)
        return output_path

    async def _convert_xlsx_to_pdf(self, xlsx_path: str, output_path: str) -> str:
        """Convert Excel file to PDF.

        Each sheet becomes a separate page with a title.

        Args:
            xlsx_path: Path to source XLSX file
            output_path: Path for output PDF file

        Returns:
            Path to the generated PDF file
        """
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)

        pdf = FPDF()
        pdf.set_margin(self.MARGIN)

        page_width = self.A4_WIDTH - 2 * self.MARGIN
        cell_height = 6  # mm
        font_size = 9

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Add new page for each sheet
            pdf.add_page()

            # Add sheet title
            pdf.set_font("Arial", "B", size=14)
            pdf.cell(0, 10, txt=f"Sheet: {sheet_name}", ln=True, align="C")
            pdf.ln(3)

            # Get data from sheet
            data = []
            max_row = 0
            for row in ws.iter_rows(values_only=True):
                if any(cell is not None for cell in row):
                    data.append([str(cell) if cell is not None else "" for cell in row])
                    max_row = max(max_row, len(data))

            if not data:
                pdf.set_font("Arial", size=10)
                pdf.cell(0, 10, txt="(Empty sheet)", ln=True)
                continue

            # Calculate column widths
            num_cols = max(len(row) for row in data) if data else 0
            if num_cols == 0:
                continue

            col_width = page_width / num_cols

            # Set font for table
            pdf.set_font("Arial", size=font_size)

            # Draw table
            for row_idx, row in enumerate(data):
                # Check if we need a new page
                if pdf.y > self.A4_HEIGHT - 20:
                    pdf.add_page()

                for col_idx, cell_value in enumerate(row):
                    # Truncate long cell content
                    max_chars = int(col_width / pdf.get_string_width("x")) - 2
                    if len(cell_value) > max_chars:
                        cell_value = cell_value[:max_chars] + "..."

                    # Draw cell border and content
                    pdf.cell(col_width, cell_height, txt=cell_value, border=1, align="L")

                pdf.ln(cell_height)

        pdf.output(output_path)
        wb.close()
        return output_path

    def _read_text_file(self, file_path: str) -> str:
        """Read text file with encoding detection.

        Args:
            file_path: Path to text file

        Returns:
            File content as string
        """
        # Try UTF-8 first
        try:
            with open(file_path, "r", encoding="utf-8") as f:
                return f.read()
        except UnicodeDecodeError:
            pass

        # Try with chardet if available
        if detect:
            with open(file_path, "rb") as f:
                raw_data = f.read()
                detected = detect(raw_data)
                encoding = detected.get("encoding", "utf-8")
                confidence = detected.get("confidence", 0)

                logger.info(f"Detected encoding: {encoding} (confidence: {confidence:.2f})")

                try:
                    return raw_data.decode(encoding)
                except (UnicodeDecodeError, LookupError):
                    pass

        # Fallback to common encodings
        encodings = ["cp1251", "latin-1"]
        for enc in encodings:
            try:
                with open(file_path, "r", encoding=enc) as f:
                    return f.read()
            except UnicodeDecodeError:
                continue

        # Last resort: latin-1 (never fails)
        with open(file_path, "r", encoding="latin-1") as f:
            return f.read()
